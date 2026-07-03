from __future__ import annotations

import argparse
import json
import math
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT_DIR / "data"
TEMPLATE_DIR = ROOT_DIR / "templates"
OUTPUT_JS = DATA_DIR / "portal-data.js"
AI_TEMPLATE = TEMPLATE_DIR / "AI产品数据填写模板.xlsx"
BEIKE_TEMPLATE = TEMPLATE_DIR / "贝壳数据填写模板.xlsx"


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync fixed Excel templates into portal-data.js")
    parser.add_argument("--output", default=str(OUTPUT_JS))
    args = parser.parse_args()

    output = Path(args.output)
    warnings: list[str] = []

    ai_file = AI_TEMPLATE if AI_TEMPLATE.exists() else None
    beike_file = BEIKE_TEMPLATE if BEIKE_TEMPLATE.exists() else None

    ai = parse_ai_workbook(ai_file, warnings) if ai_file else empty_ai("未找到 AI 产品数据填写模板")
    if not ai_file:
        warnings.append("未找到 templates/AI产品数据填写模板.xlsx")

    beike = parse_beike_workbook(beike_file, warnings) if beike_file else empty_beike("未找到贝壳数据填写模板")
    if not beike_file:
        warnings.append("未找到 templates/贝壳数据填写模板.xlsx")

    payload = build_payload(ai, beike, warnings)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        "window.HT_PORTAL_REAL_DATA = "
        + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )

    print(json.dumps({
        "ok": True,
        "output": str(output),
        "aiTemplate": str(ai_file) if ai_file else None,
        "beikeTemplate": str(beike_file) if beike_file else None,
        "warnings": warnings,
    }, ensure_ascii=False, indent=2))
    return 0


def parse_ai_workbook(path: Path, warnings: list[str]) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = get_sheet(wb, "ChatGPT美国DAU数据", 1)
    if not ws:
        warnings.append(f"{path.name} 缺少模板页 ChatGPT美国DAU数据")
        return empty_ai(f"{path.name} 缺少 AI 模板数据页")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        warnings.append(f"{path.name} 没有可读取数据")
        return empty_ai(f"{path.name} 为空")

    long_header_row = find_header_row(rows, ["Date", "Product", "Region", "Metric", "Value"])
    records = parse_ai_long(rows, long_header_row) if long_header_row is not None else parse_ai_wide(rows)
    records = sorted(records, key=lambda item: (item["date"], item["product"], item["region"], item["metric"]))
    attach_group_changes(records, ["product", "region", "metric"])

    latest = latest_by(records, ["product", "region", "metric"])
    ai = {
        "sourceFile": path.name,
        "records": records,
        "latest": latest,
        "products": sorted({item["product"] for item in records}),
        "regions": sorted({item["region"] for item in records}),
        "metrics": sorted({item["metric"] for item in records}),
        "dateRange": date_range(records),
    }
    ai["blocks"] = build_ai_blocks(ai)
    ai["kpis"] = build_ai_kpis(ai)
    ai["briefs"] = build_ai_briefs(ai)
    return ai


def parse_ai_long(rows: list[tuple[Any, ...]], header_idx: int) -> list[dict[str, Any]]:
    headers = [clean_text(value) for value in rows[header_idx]]
    idx = {name: headers.index(name) for name in headers if name}
    records = []
    for row in rows[header_idx + 1:]:
        date_iso = to_iso_date(row_value(row, idx.get("Date")))
        product = clean_text(row_value(row, idx.get("Product")))
        region = normalize_region(clean_text(row_value(row, idx.get("Region"))))
        metric = normalize_ai_metric(clean_text(row_value(row, idx.get("Metric"))))
        value = to_number(row_value(row, idx.get("Value")))
        if not (date_iso and product and region and metric and value is not None):
            continue
        records.append({
            "date": date_iso,
            "product": product,
            "region": region,
            "metric": metric,
            "value": value,
            "unit": clean_text(row_value(row, idx.get("Unit"))) or default_ai_unit(metric),
            "source": clean_text(row_value(row, idx.get("Source"))),
            "notes": clean_text(row_value(row, idx.get("Notes"))),
        })
    return records


def parse_ai_wide(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    header1 = rows[0] if len(rows) > 0 else ()
    header2 = rows[1] if len(rows) > 1 else ()
    col_specs = []
    current_product = ""
    max_cols = max(len(header1), len(header2))
    for idx in range(1, max_cols):
        h1 = clean_text(row_value(header1, idx))
        h2 = clean_text(row_value(header2, idx))
        if h1:
            current_product = h1
        if not h2 or not current_product:
            continue
        region = "US" if "US" in h2.upper() else "Global" if "GLOBAL" in h2.upper() else ""
        metric = normalize_ai_metric(h2)
        if not region or not metric:
            continue
        col_specs.append({
            "idx": idx,
            "product": current_product,
            "region": region,
            "metric": metric,
            "unit": default_ai_unit(metric),
        })

    records = []
    for row in rows[2:]:
        date_iso = to_iso_date(row_value(row, 0))
        if not date_iso:
            continue
        for spec in col_specs:
            value = to_number(row_value(row, spec["idx"]))
            if value is None:
                continue
            records.append({
                "date": date_iso,
                "product": spec["product"],
                "region": spec["region"],
                "metric": spec["metric"],
                "value": value,
                "unit": spec["unit"],
                "source": "",
                "notes": "",
            })
    return records


def build_ai_blocks(ai: dict[str, Any]) -> dict[str, Any]:
    return {
        "dau": build_ai_metric_block(ai.get("records", []), "DAU", "AI 产品 DAU", "百万"),
        "avgTime": build_ai_metric_block(ai.get("records", []), "AvgTime", "AI 产品人均使用时长", "分钟"),
    }


def build_ai_metric_block(records: list[dict[str, Any]], metric: str, title: str, unit: str) -> dict[str, Any]:
    metric_records = [item for item in records if item.get("metric") == metric]
    products = sort_preferred({item["product"] for item in metric_records}, ["ChatGPT", "Gemini"])
    regions = sort_preferred({item["region"] for item in metric_records}, ["US", "Global"])
    columns = ["日期"] + [f"{product} {display_region(region)}" for product in products for region in regions]
    by_key = {
        (item["date"], item["product"], item["region"]): item["value"]
        for item in metric_records
    }
    rows = []
    for date_iso in sorted({item["date"] for item in metric_records}):
        row = {"日期": date_iso}
        for product in products:
            for region in regions:
                row[f"{product} {display_region(region)}"] = by_key.get((date_iso, product, region))
        rows.append(row)
    return {
        "id": metric,
        "title": title,
        "unit": unit,
        "columns": columns,
        "rows": rows,
        "dateRange": date_range(metric_records),
    }


def parse_beike_workbook(path: Path, warnings: list[str]) -> dict[str, Any]:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheetnames = list(wb.sheetnames)
    core_ws = get_sheet(wb, "QM核心App数据", 1)
    beike_ws = get_sheet(wb, "QM-贝壳找房", 2)

    core_records: list[dict[str, Any]] = []
    city_records: list[dict[str, Any]] = []
    yearly_records: list[dict[str, Any]] = []
    blocks: dict[str, Any] = {
        "coreWau": empty_block("coreWau", "核心 App WAU", "App"),
        "coreDuration": empty_block("coreDuration", "核心 App 使用总时长", "App"),
        "cityWau": empty_block("cityWau", "贝壳找房城市 WAU", "城市"),
        "yearlyWau": empty_block("yearlyWau", "贝壳找房历年 WAU", "年份"),
        "yearlyAvgTime": empty_block("yearlyAvgTime", "贝壳找房历年人均单日使用时长", "年份"),
    }

    if core_ws:
        rows = list(core_ws.iter_rows(values_only=True))
        long_header_row = find_header_row(rows, ["Date", "App", "Metric", "Value"])
        if long_header_row is not None:
            core_records = parse_beike_long(rows, long_header_row, default_dimension="核心App")
            blocks["coreWau"] = records_to_entity_block(core_records, "WAU", "核心 App WAU", "App")
            blocks["coreDuration"] = records_to_entity_block(core_records, "Duration", "核心 App 使用总时长", "App")
        else:
            sections = [(idx, normalize_beike_metric(clean_text(row_value(row, 0)))) for idx, row in enumerate(rows)]
            sections = [(idx, metric) for idx, metric in sections if metric]
            if len(sections) >= 1:
                blocks["coreWau"], wau_records = parse_wide_block(
                    rows, sections[0][0], "coreWau", "核心 App WAU", "App", "WAU", "万人", "核心App"
                )
                core_records.extend(wau_records)
            if len(sections) >= 2:
                blocks["coreDuration"], duration_records = parse_wide_block(
                    rows, sections[1][0], "coreDuration", "核心 App 使用总时长", "App", "Duration", "万分钟", "核心App"
                )
                core_records.extend(duration_records)
    else:
        warnings.append(f"{path.name} 缺少模板页 QM核心App数据；当前 sheets={sheetnames}")

    if beike_ws:
        rows = list(beike_ws.iter_rows(values_only=True))
        long_header_row = find_header_row(rows, ["Date", "App", "City", "Metric", "Value"])
        if long_header_row is not None:
            city_records = parse_beike_long(rows, long_header_row, default_dimension="城市")
            blocks["cityWau"] = records_to_entity_block(city_records, "WAU", "贝壳找房城市 WAU", "城市")
        else:
            sections = [(idx, normalize_beike_metric(clean_text(row_value(row, 0)))) for idx, row in enumerate(rows)]
            sections = [(idx, metric) for idx, metric in sections if metric]
            if len(sections) >= 1:
                blocks["cityWau"], city_records = parse_wide_block(
                    rows, sections[0][0], "cityWau", "贝壳找房城市 WAU", "城市", "WAU", "万人", "城市",
                    fixed_app="贝壳找房",
                )
            if len(sections) >= 2:
                blocks["yearlyWau"], yearly_wau_records = parse_wide_block(
                    rows, sections[1][0], "yearlyWau", "贝壳找房历年 WAU", "年份", "WAU", "万人", "年度",
                    fixed_app="贝壳找房",
                )
                yearly_records.extend(yearly_wau_records)
            if len(sections) >= 3:
                blocks["yearlyAvgTime"], yearly_time_records = parse_wide_block(
                    rows, sections[2][0], "yearlyAvgTime", "贝壳找房历年人均单日使用时长", "年份", "AvgTime", "分钟", "年度",
                    fixed_app="贝壳找房",
                )
                yearly_records.extend(yearly_time_records)
    else:
        warnings.append(f"{path.name} 缺少模板页 QM-贝壳找房；当前 sheets={sheetnames}")

    core_records = sorted(core_records, key=lambda item: (item["date"], item["app"], item["metric"]))
    city_records = sorted(city_records, key=lambda item: (item["date"], item.get("city", ""), item["metric"]))
    yearly_records = sorted(yearly_records, key=lambda item: (item["metric"], str(item.get("year", "")), item["date"]))
    attach_group_changes(core_records, ["app", "metric"])
    attach_group_changes(city_records, ["city", "metric"])

    beike = {
        "sourceFile": path.name,
        "coreRecords": core_records,
        "cityRecords": city_records,
        "yearlyRecords": yearly_records,
        "coreLatest": latest_by(core_records, ["app", "metric"]),
        "cityLatest": latest_by(city_records, ["city", "metric"]),
        "yearlyLatest": latest_by(yearly_records, ["year", "metric"]),
        "apps": sorted({item["app"] for item in core_records}),
        "cities": sorted({item["city"] for item in city_records if item.get("city")}),
        "years": sorted({item["year"] for item in yearly_records if item.get("year")}),
        "metrics": sorted({item["metric"] for item in [*core_records, *city_records, *yearly_records]}),
        "dateRange": date_range([*core_records, *city_records, *yearly_records]),
        "blocks": blocks,
    }
    beike["kpis"] = build_beike_kpis(beike)
    beike["briefs"] = build_beike_briefs(beike)
    return beike


def get_sheet(wb, name: str, fallback_index: int):
    if name in wb.sheetnames:
        return wb[name]
    if len(wb.worksheets) > fallback_index:
        return wb.worksheets[fallback_index]
    return None


def parse_beike_long(rows: list[tuple[Any, ...]], header_idx: int, default_dimension: str) -> list[dict[str, Any]]:
    headers = [clean_text(value) for value in rows[header_idx]]
    idx = {name: headers.index(name) for name in headers if name}
    records = []
    for row in rows[header_idx + 1:]:
        date_iso = to_iso_date(row_value(row, idx.get("Date")))
        app = clean_text(row_value(row, idx.get("App")))
        metric = normalize_beike_metric(clean_text(row_value(row, idx.get("Metric"))))
        value = to_number(row_value(row, idx.get("Value")))
        if not (date_iso and app and metric and value is not None):
            continue
        city = clean_label(clean_text(row_value(row, idx.get("City"))))
        year = clean_text(row_value(row, idx.get("Year")))
        records.append({
            "date": date_iso,
            "app": app,
            "city": city,
            "year": year,
            "dimension": "城市" if city else "年度" if year else default_dimension,
            "metric": metric,
            "value": value,
            "unit": clean_text(row_value(row, idx.get("Unit"))) or default_beike_unit(metric),
            "wow": to_number(row_value(row, idx.get("WoW"))),
            "yoy": to_number(row_value(row, idx.get("YoY"))),
            "source": clean_text(row_value(row, idx.get("Source"))),
            "notes": clean_text(row_value(row, idx.get("Notes"))),
        })
    return records


def parse_wide_block(
    rows: list[tuple[Any, ...]],
    section_idx: int,
    block_id: str,
    title: str,
    label_header: str,
    metric: str,
    unit: str,
    dimension: str,
    fixed_app: str | None = None,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    header = rows[section_idx]
    date_cols, wow_col, yoy_col = find_date_columns(header)
    columns = [label_header] + [date_iso for _, date_iso in date_cols]
    if wow_col is not None:
        columns.append("WoW")
    if yoy_col is not None:
        columns.append("YoY")

    block_rows = []
    records = []
    row_idx = section_idx + 1
    while row_idx < len(rows):
        row = rows[row_idx]
        label_value = row_value(row, 0)
        label = clean_label(clean_text(label_value))
        if not label:
            break
        block_row = {label_header: label}
        wow = to_number(row_value(row, wow_col))
        yoy = to_number(row_value(row, yoy_col))
        for col_idx, date_iso in date_cols:
            value = to_number(row_value(row, col_idx))
            block_row[date_iso] = value
            if value is None:
                continue
            record = {
                "date": date_iso,
                "app": fixed_app or label,
                "city": label if dimension == "城市" else "",
                "year": label if dimension == "年度" else "",
                "dimension": dimension,
                "metric": metric,
                "value": value,
                "unit": unit,
                "wow": wow,
                "yoy": yoy,
                "source": "QuestMobile",
                "notes": "",
            }
            records.append(record)
        if wow_col is not None:
            block_row["WoW"] = wow
        if yoy_col is not None:
            block_row["YoY"] = yoy
        block_rows.append(block_row)
        row_idx += 1

    return {
        "id": block_id,
        "title": title,
        "label": label_header,
        "metric": metric,
        "unit": unit,
        "columns": columns,
        "dateColumns": [date_iso for _, date_iso in date_cols],
        "rows": block_rows,
    }, records


def records_to_entity_block(records: list[dict[str, Any]], metric: str, title: str, label_header: str) -> dict[str, Any]:
    metric_records = [item for item in records if item["metric"] == metric]
    label_key = "city" if label_header == "城市" else "year" if label_header == "年份" else "app"
    labels = sorted({item.get(label_key, "") for item in metric_records if item.get(label_key, "")})
    dates = sorted({item["date"] for item in metric_records})
    by_key = {(item.get(label_key, ""), item["date"]): item["value"] for item in metric_records}
    rows = []
    for label in labels:
        row = {label_header: label}
        for date_iso in dates:
            row[date_iso] = by_key.get((label, date_iso))
        rows.append(row)
    return {
        "id": metric,
        "title": title,
        "label": label_header,
        "metric": metric,
        "unit": default_beike_unit(metric),
        "columns": [label_header] + dates,
        "dateColumns": dates,
        "rows": rows,
    }


def empty_block(block_id: str, title: str, label: str) -> dict[str, Any]:
    return {
        "id": block_id,
        "title": title,
        "label": label,
        "metric": "",
        "unit": "",
        "columns": [label],
        "dateColumns": [],
        "rows": [],
    }


def build_payload(ai: dict[str, Any], beike: dict[str, Any], warnings: list[str]) -> dict[str, Any]:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {
        "meta": {
            "version": f"ExcelSync-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
            "lastUpdated": now,
            "source": "本地 Excel 填写模板",
            "warnings": warnings,
            "files": {
                "ai": ai.get("sourceFile", ""),
                "beike": beike.get("sourceFile", ""),
            },
        },
        "kpis": [*ai.get("kpis", []), *beike.get("kpis", [])][:8],
        "dailyBriefs": [*ai.get("briefs", []), *beike.get("briefs", [])],
        "ai": ai,
        "beike": beike,
    }


def build_ai_kpis(ai: dict[str, Any]) -> list[dict[str, Any]]:
    latest = ai.get("latest", [])
    specs = [
        ("ChatGPT 全球 DAU", {"product": "ChatGPT", "region": "Global", "metric": "DAU"}, "百万"),
        ("ChatGPT 美国 DAU", {"product": "ChatGPT", "region": "US", "metric": "DAU"}, "百万"),
        ("Gemini 全球 DAU", {"product": "Gemini", "region": "Global", "metric": "DAU"}, "百万"),
        ("Gemini 美国 DAU", {"product": "Gemini", "region": "US", "metric": "DAU"}, "百万"),
    ]
    kpis = [kpi_from_record(label, record, unit) for label, criteria, unit in specs if (record := first_match(latest, **criteria))]
    if not kpis:
        kpis.append({"label": "AI 产品数据", "value": "暂无", "unit": "", "change": None, "note": "请检查 templates 文件夹"})
    return kpis[:4]


def build_beike_kpis(beike: dict[str, Any]) -> list[dict[str, Any]]:
    latest = beike.get("coreLatest", [])
    specs = [
        ("贝壳找房 WAU", {"app": "贝壳找房", "metric": "WAU"}, "万人"),
        ("链家 WAU", {"app": "链家", "metric": "WAU"}, "万人"),
        ("贝壳租房 WAU", {"app": "贝壳租房", "metric": "WAU"}, "万人"),
        ("贝壳找房使用总时长", {"app": "贝壳找房", "metric": "Duration"}, "万分钟"),
    ]
    kpis = [kpi_from_record(label, record, unit) for label, criteria, unit in specs if (record := first_match(latest, **criteria))]
    if not kpis:
        kpis.append({"label": "贝壳数据", "value": "暂无", "unit": "", "change": None, "note": "请检查 templates 文件夹"})
    return kpis[:4]


def build_ai_briefs(ai: dict[str, Any]) -> list[dict[str, str]]:
    date_info = ai.get("dateRange", {})
    return [
        {
            "tag": "AI 数据",
            "title": ai.get("sourceFile", "AI 数据未找到"),
            "note": f"样本区间 {date_info.get('start', '--')} 至 {date_info.get('end', '--')}，共 {len(ai.get('records', []))} 条观测。",
        },
        {
            "tag": "AI 指标",
            "title": "DAU / 人均使用时长",
            "note": "指标明细按 DAU 和人均使用时长分项展示。",
        },
    ]


def build_beike_briefs(beike: dict[str, Any]) -> list[dict[str, str]]:
    blocks = beike.get("blocks", {})
    block_titles = [block.get("title", "") for block in blocks.values() if block.get("rows")]
    return [
        {
            "tag": "贝壳数据",
            "title": beike.get("sourceFile", "贝壳数据未找到"),
            "note": f"核心 App {len(beike.get('coreRecords', []))} 条，城市 {len(beike.get('cityRecords', []))} 条，年度序列 {len(beike.get('yearlyRecords', []))} 条。",
        },
        {
            "tag": "贝壳指标",
            "title": " / ".join(block_titles[:3]) or "暂无指标明细",
            "note": f"共 {len(block_titles)} 个分项指标，按指标口径独立展示。",
        },
    ]


def kpi_from_record(label: str, record: dict[str, Any], fallback_unit: str) -> dict[str, Any]:
    change = record.get("changePct")
    return {
        "label": label,
        "value": round(record["value"], 2),
        "unit": record.get("unit") or fallback_unit,
        "change": round(change, 2) if change is not None else None,
        "note": f"截至日期 {record['date']}",
    }


def latest_by(records: list[dict[str, Any]], keys: list[str]) -> list[dict[str, Any]]:
    latest: dict[tuple[Any, ...], dict[str, Any]] = {}
    for record in sorted(records, key=lambda item: item["date"]):
        key = tuple(record.get(name, "") for name in keys)
        latest[key] = record
    return sorted((dict(item) for item in latest.values()), key=lambda item: tuple(str(item.get(name, "")) for name in keys))


def attach_group_changes(records: list[dict[str, Any]], keys: list[str]) -> None:
    grouped: dict[tuple[Any, ...], list[dict[str, Any]]] = {}
    for record in records:
        key = tuple(record.get(name, "") for name in keys)
        grouped.setdefault(key, []).append(record)
    for group_records in grouped.values():
        group_records.sort(key=lambda item: item["date"])
        previous = None
        for record in group_records:
            if previous is not None and previous.get("value") not in (None, 0):
                record["changePct"] = (record["value"] / previous["value"] - 1) * 100
            else:
                record["changePct"] = None
            previous = record


def date_range(records: list[dict[str, Any]]) -> dict[str, str]:
    dates = sorted({item["date"] for item in records if item.get("date")})
    return {"start": dates[0], "end": dates[-1]} if dates else {"start": "", "end": ""}


def find_header_row(rows: list[tuple[Any, ...]], required: list[str]) -> int | None:
    required_set = {item.lower() for item in required}
    for idx, row in enumerate(rows[:12]):
        values = {clean_text(value).lower() for value in row if clean_text(value)}
        if required_set.issubset(values):
            return idx
    return None


def find_date_columns(row: tuple[Any, ...]) -> tuple[list[tuple[int, str]], int | None, int | None]:
    date_cols = []
    wow_col = None
    yoy_col = None
    for idx, value in enumerate(row):
        text = clean_text(value)
        if text.upper() == "WOW":
            wow_col = idx
            continue
        if text.upper() == "YOY":
            yoy_col = idx
            continue
        date_iso = to_iso_date(value)
        if date_iso:
            date_cols.append((idx, date_iso))
    return date_cols, wow_col, yoy_col


def first_match(records: Iterable[dict[str, Any]], **criteria: Any) -> dict[str, Any] | None:
    for record in records:
        if all(record.get(key) == value for key, value in criteria.items()):
            return record
    return None


def row_value(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def clean_label(value: str) -> str:
    value = re.sub(r"[（(]右轴[）)]", "", value)
    return value.strip()


def normalize_region(value: str) -> str:
    text = value.lower()
    if text in {"us", "usa", "美国"}:
        return "US"
    if text in {"global", "全球"}:
        return "Global"
    return value


def display_region(value: str) -> str:
    return {"US": "美国", "Global": "全球"}.get(value, value)


def normalize_ai_metric(value: str) -> str:
    text = value.lower()
    if "dau" in text:
        return "DAU"
    if "时长" in value or "time" in text or "avg" in text:
        return "AvgTime"
    return value if value in {"DAU", "AvgTime"} else ""


def default_ai_unit(metric: str) -> str:
    return "百万" if metric == "DAU" else "分钟" if metric == "AvgTime" else ""


def normalize_beike_metric(value: str) -> str:
    text = value.lower()
    if "wau" in text:
        return "WAU"
    if "使用时长" in value or "duration" in text:
        return "Duration"
    if "人均" in value or "avgtime" in text:
        return "AvgTime"
    if value in {"WAU", "Duration", "AvgTime"}:
        return value
    return ""


def default_beike_unit(metric: str) -> str:
    return {"WAU": "万人", "Duration": "万分钟", "AvgTime": "分钟"}.get(metric, "")


def to_iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and value > 20000:
        try:
            return from_excel(value).strftime("%Y-%m-%d")
        except Exception:
            return ""
    if isinstance(value, str):
        text = value.strip()
        if not text or text.upper() in {"WOW", "YOY"}:
            return ""
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(text[:10], fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return ""


def to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if math.isnan(value) or math.isinf(value):
            return None
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text in {"-", "--"}:
            return None
        is_percent = text.endswith("%")
        if is_percent:
            text = text[:-1]
        try:
            number = float(text)
            return number / 100 if is_percent else number
        except ValueError:
            return None
    return None


def sort_preferred(values: Iterable[str], preferred: list[str]) -> list[str]:
    unique = list(values)
    rank = {value: idx for idx, value in enumerate(preferred)}
    return sorted(unique, key=lambda value: (rank.get(value, len(rank)), value))


def empty_ai(message: str) -> dict[str, Any]:
    return {
        "sourceFile": "",
        "records": [],
        "latest": [],
        "products": [],
        "regions": [],
        "metrics": [],
        "dateRange": {"start": "", "end": ""},
        "blocks": {
            "dau": empty_block("dau", "AI 产品 DAU", "日期"),
            "avgTime": empty_block("avgTime", "AI 产品人均时长", "日期"),
        },
        "kpis": [{"label": "AI 产品数据", "value": "暂无", "unit": "", "change": None, "note": message}],
        "briefs": [],
        "message": message,
    }


def empty_beike(message: str) -> dict[str, Any]:
    return {
        "sourceFile": "",
        "coreRecords": [],
        "cityRecords": [],
        "yearlyRecords": [],
        "coreLatest": [],
        "cityLatest": [],
        "yearlyLatest": [],
        "apps": [],
        "cities": [],
        "years": [],
        "metrics": [],
        "dateRange": {"start": "", "end": ""},
        "blocks": {
            "coreWau": empty_block("coreWau", "核心 App WAU", "App"),
            "coreDuration": empty_block("coreDuration", "核心 App 使用总时长", "App"),
            "cityWau": empty_block("cityWau", "贝壳找房城市 WAU", "城市"),
            "yearlyWau": empty_block("yearlyWau", "贝壳找房历年 WAU", "年份"),
            "yearlyAvgTime": empty_block("yearlyAvgTime", "贝壳找房历年人均单日使用时长", "年份"),
        },
        "kpis": [{"label": "贝壳数据", "value": "暂无", "unit": "", "change": None, "note": message}],
        "briefs": [],
        "message": message,
    }


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        raise
